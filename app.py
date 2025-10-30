# Sistema_Pedidosv12.py
import os
import shutil
import logging
import traceback
from decimal import Decimal
from datetime import datetime, timedelta
from functools import wraps
from pathlib import Path
from urllib.parse import urlparse, ParseResult, urlunparse, quote_plus

from flask import (
    Flask,
    abort,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import load_workbook
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy.exc import IntegrityError
from sqlalchemy import func, text
from sqlalchemy.orm import selectinload

# =====================================================
# CONFIGURACAO DE PASTAS E BANCO
# =====================================================
ROOT_DIR = Path(__file__).resolve().parent


def _normalize_db_url(url: str | None) -> str | None:
    if not url:
        return None
    # Aceita postgres:// e transforma em postgresql:// para compatibilidade
    if url.startswith("postgres://"):
        return url.replace("postgres://", "postgresql://", 1)
    return url


def _build_postgres_url_from_parts() -> str | None:
    user = os.environ.get("POSTGRES_USER") or os.environ.get("PGUSER")
    password = os.environ.get("POSTGRES_PASSWORD") or os.environ.get("PGPASSWORD")
    db = os.environ.get("POSTGRES_DB") or os.environ.get("PGDATABASE")
    host = os.environ.get("POSTGRES_HOST") or os.environ.get("PGHOST")
    port = os.environ.get("POSTGRES_PORT") or os.environ.get("PGPORT")
    sslmode = os.environ.get("POSTGRES_SSLMODE") or os.environ.get("PGSSLMODE")
    if not (user and password and db and host):
        return None

    user_quoted = quote_plus(user)
    password_quoted = quote_plus(password)
    netloc = f"{user_quoted}:{password_quoted}@{host}"
    if port:
        netloc = f"{netloc}:{port}"
    path = f"/{db}"
    query = ""
    if sslmode:
        query = f"sslmode={sslmode}"
    parsed = ParseResult(scheme="postgresql", netloc=netloc, path=path, params="", query=query, fragment="")
    return urlunparse(parsed)


STORAGE_ROOT = Path(
    os.environ.get("PEDIDOS_STORAGE_DIR")
    or os.environ.get("PEDIDOS_BASE_DIR")
    or ROOT_DIR
).expanduser()
STORAGE_ROOT.mkdir(parents=True, exist_ok=True)


def _default_sqlite_path() -> Path:
    db_path = Path(
        os.environ.get("PEDIDOS_DB_PATH") or (STORAGE_ROOT / "pedidos.db")
    ).expanduser()
    if not db_path.exists():
        legacy_db = ROOT_DIR.parent / "pedidos.db"
        if legacy_db.exists():
            db_path = legacy_db
    db_path.parent.mkdir(parents=True, exist_ok=True)
    return db_path


# ---------- DADOS DO SEU POSTGRES (fallback)
# Essas informações vieram do seu Render — usadas apenas se DATABASE_URL não estiver definida.
RENDER_EXTERNAL_DB = (
    "postgresql://pedidos_db_ihvt_user:M8il2h2WH7OxTCvQ5FQkKMxylVuEzPhd"
    "@dpg-d416uiqli9vc739ftdbg-a.oregon-postgres.render.com/pedidos_db_ihvt"
)

# Prefer DATABASE_URL em variáveis de ambiente (Render geralmente fornece DATABASE_URL).
ENV_DB_URL = os.environ.get("DATABASE_URL") or os.environ.get("RENDER_DATABASE_URL") or os.environ.get("POSTGRES_URL")
DB_URL = _normalize_db_url(ENV_DB_URL) if ENV_DB_URL else None

force_sqlite_env = os.environ.get("PEDIDOS_FORCE_SQLITE", "")
FORCE_SQLITE = force_sqlite_env.strip().lower() in {"1", "true", "yes", "on"}

if not DB_URL and not FORCE_SQLITE:
    # tenta construir a partir de partes (se definidas)
    built = _build_postgres_url_from_parts()
    if built:
        DB_URL = built

# Se ainda nǜa existir, usar o fallback com suas credenciais (mas preferĕvel definir DATABASE_URL no painel).
if not DB_URL and not FORCE_SQLITE:
    DB_URL = _normalize_db_url(RENDER_EXTERNAL_DB)

# Se DB_URL ainda for None (improvǕvel) ou houver forǐ3o para sqlite, cai para sqlite local.
if not DB_URL or FORCE_SQLITE:
    _sqlite_path = _default_sqlite_path()
    DATABASE_URI = f"sqlite:///{_sqlite_path}"
else:
    DATABASE_URI = DB_URL

modelo_default = ROOT_DIR / "modelo_pedido.xlsm"
MODELO_PATH = Path(os.environ.get("PEDIDOS_MODELO_PATH") or modelo_default)
if not MODELO_PATH.exists():
    alt_modelo = STORAGE_ROOT / "modelo_pedido.xlsm"
    if alt_modelo.exists():
        MODELO_PATH = alt_modelo
    else:
        legacy_modelo = ROOT_DIR.parent / "modelo_pedido.xlsm"
        if legacy_modelo.exists():
            MODELO_PATH = legacy_modelo

PASTA_PEDIDOS_GERADOS = Path(
    os.environ.get("PEDIDOS_GERADOS_DIR") or (STORAGE_ROOT / "Pedidos Gerados")
)
PASTA_PEDIDOS_APROVADOS = Path(
    os.environ.get("PEDIDOS_APROVADOS_DIR") or (STORAGE_ROOT / "PedidosAprovados")
)

if not os.environ.get("PEDIDOS_GERADOS_DIR"):
    legacy_gerados = ROOT_DIR.parent / "Pedidos Gerados"
    if legacy_gerados.exists():
        PASTA_PEDIDOS_GERADOS = legacy_gerados

if not os.environ.get("PEDIDOS_APROVADOS_DIR"):
    legacy_aprovados = ROOT_DIR.parent / "PedidosAprovados"
    if legacy_aprovados.exists():
        PASTA_PEDIDOS_APROVADOS = legacy_aprovados

PASTA_PEDIDOS_GERADOS.mkdir(parents=True, exist_ok=True)
PASTA_PEDIDOS_APROVADOS.mkdir(parents=True, exist_ok=True)


# =====================================================
# APP / LOGGING / DB
# =====================================================
app = Flask(__name__, template_folder="templates", static_folder="static")
secret_key = os.environ.get("PEDIDOS_SECRET_KEY", "change-me")
app.config["SECRET_KEY"] = secret_key
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(hours=8)

app.config["SQLALCHEMY_DATABASE_URI"] = DATABASE_URI
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

# Se for PostgreSQL e não tiver sslmode na query, passa connect_args sslmode=require
try:
    parsed = urlparse(DATABASE_URI)
    scheme = parsed.scheme or ""
    if scheme.startswith("postgres") or scheme.startswith("postgresql"):
        if "sslmode=" not in (parsed.query or ""):
            app.config.setdefault("SQLALCHEMY_ENGINE_OPTIONS", {})
            engine_opts = app.config["SQLALCHEMY_ENGINE_OPTIONS"]
            if "connect_args" not in engine_opts:
                engine_opts["connect_args"] = {"sslmode": "require"}
            else:
                engine_opts["connect_args"].setdefault("sslmode", "require")
            app.config["SQLALCHEMY_ENGINE_OPTIONS"] = engine_opts
except Exception:
    pass

db = SQLAlchemy(app)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("pedidos_app")


# =====================================================
# MODELOS
# =====================================================
class User(db.Model):
    __tablename__ = "users"
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(150), unique=True, nullable=False, index=True)
    password = db.Column(db.String(300), nullable=False)
    role = db.Column(db.String(50), nullable=False, default="creator")


class Fornecedor(db.Model):
    __tablename__ = "fornecedores"
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(255), unique=True, nullable=False)


class Pedido(db.Model):
    __tablename__ = "pedidos"
    id = db.Column(db.Integer, primary_key=True)
    fornecedor = db.Column(db.String(255), nullable=False)
    arquivo_excel = db.Column(db.String(500), nullable=True)
    arquivo_pdf = db.Column(db.String(500), nullable=True)
    created_at = db.Column(db.DateTime(timezone=False), default=datetime.utcnow, nullable=False)
    status = db.Column(db.String(50), default="Pendente", nullable=False)
    created_by = db.Column(db.String(150), nullable=True)

    itens = db.relationship(
        "PedidoItem",
        backref="pedido",
        cascade="all, delete-orphan",
        lazy="selectin",
    )


class PedidoItem(db.Model):
    __tablename__ = "pedidos_items"
    id = db.Column(db.Integer, primary_key=True)
    pedido_id = db.Column(
        db.Integer,
        db.ForeignKey("pedidos.id", ondelete="CASCADE"),
        nullable=False,
        index=True,
    )
    codigo = db.Column(db.String(200), nullable=False)
    descricao = db.Column(db.Text, nullable=False)
    quantidade = db.Column(db.Integer, nullable=False)
    prefixo = db.Column(db.String(50), nullable=True)
    valor = db.Column(db.Numeric(12, 2), nullable=False)
    estoque = db.Column(db.Integer, nullable=True)


# =====================================================
# UTILITARIOS / AUTH
# =====================================================
def current_user():
    return session.get("user")


def login_required(view):
    @wraps(view)
    def wrapper(*args, **kwargs):
        if "user" not in session:
            next_url = request.path if request.method == "GET" else None
            return redirect(url_for("login", next=next_url))
        return view(*args, **kwargs)

    return wrapper


def api_login_required(view):
    @wraps(view)
    def wrapper(*args, **kwargs):
        if "user" not in session:
            return jsonify({"error": "autenticacao requerida"}), 401
        return view(*args, **kwargs)

    return wrapper


def roles_required(*roles):
    def decorator(view):
        @wraps(view)
        def wrapper(*args, **kwargs):
            user = current_user()
            if not user:
                return jsonify({"error": "autenticacao requerida"}), 401
            if user.get("role") not in roles:
                return jsonify({"error": "permissao negada"}), 403
            return view(*args, **kwargs)

        return wrapper

    return decorator


def _is_password_hashed(value: str | None) -> bool:
    value = (value or "").strip()
    if not value:
        return False
    known_prefixes = ("pbkdf2:", "scrypt:", "sha256$", "sha1$")
    return value.startswith(known_prefixes)


def _serialize_user(user: User | None) -> dict | None:
    if not user:
        return None
    return {"id": user.id, "username": user.username, "role": user.role}


# =====================================================
# USUARIOS
# =====================================================
def garantir_usuarios_iniciais():
    # SENHA PADRAO "1234" conforme solicitado
    defaults = [
        ("MIGUEL", "1234", "creator"),
        ("MICHEL", "1234", "approver"),
        ("LUCAS", "1234", "admin"),
    ]
    for username, password, role in defaults:
        user = User.query.filter(func.upper(User.username) == username).first()
        hashed = generate_password_hash(password)
        if user:
            needs_update = False
            if not _is_password_hashed(user.password):
                user.password = hashed
                needs_update = True
            if user.role != role:
                user.role = role
                needs_update = True
            if needs_update:
                db.session.add(user)
        else:
            db.session.add(User(username=username, password=hashed, role=role))
    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        logger.exception("Erro ao garantir usuarios iniciais")


def verificar_login(username: str, password: str) -> dict | None:
    username = (username or "").strip().upper()
    password = (password or "").strip()
    if not username or not password:
        return None
    user = User.query.filter(func.upper(User.username) == username).first()
    if not user:
        return None
    stored = user.password or ""
    if stored:
        try:
            if check_password_hash(stored, password):
                return _serialize_user(user)
        except (TypeError, ValueError):
            pass
    if stored == password:
        user.password = generate_password_hash(password)
        try:
            db.session.commit()
        except Exception:
            db.session.rollback()
        return _serialize_user(user)
    return None


def change_password(username: str, new_password: str):
    username = (username or "").strip().upper()
    if not username:
        raise ValueError("Usuario nao encontrado")
    user = User.query.filter(func.upper(User.username) == username).first()
    if not user:
        raise ValueError("Usuario nao encontrado")
    user.password = generate_password_hash(new_password)
    db.session.add(user)
    db.session.commit()


def create_user(username: str, password: str, role: str):
    username = (username or "").strip().upper()
    password = (password or "").strip()
    role = (role or "creator").strip()
    if not username or not password:
        raise ValueError("Usuario e senha devem ser informados")
    hashed = generate_password_hash(password)
    user = User(username=username, password=hashed, role=role)
    db.session.add(user)
    try:
        db.session.commit()
    except IntegrityError as exc:
        db.session.rollback()
        raise ValueError("Usuario ja existe") from exc


def delete_user(username: str):
    username = (username or "").strip().upper()
    if not username:
        raise ValueError("Usuario nao encontrado")
    user = User.query.filter(func.upper(User.username) == username).first()
    if not user:
        raise ValueError("Usuario nao encontrado")
    db.session.delete(user)
    db.session.commit()


def list_users() -> list[dict]:
    users = User.query.order_by(User.username).all()
    return [_serialize_user(user) for user in users if user]


# =====================================================
# FORNECEDORES
# =====================================================
def add_supplier(nome: str):
    nome = (nome or "").strip()
    if not nome:
        raise ValueError("Informe o nome do fornecedor")
    fornecedor = Fornecedor(nome=nome)
    db.session.add(fornecedor)
    try:
        db.session.commit()
    except IntegrityError as exc:
        db.session.rollback()
        raise ValueError("Fornecedor ja existe") from exc
    return nome


def list_suppliers() -> list[str]:
    fornecedores = Fornecedor.query.order_by(Fornecedor.nome).all()
    return [f.nome for f in fornecedores]


# =====================================================
# PEDIDOS
# =====================================================
def purge_old_pedidos(days: int = 135) -> int:
    cutoff = datetime.utcnow() - timedelta(days=days)
    try:
        removed = (
            db.session.query(Pedido)
            .filter(Pedido.created_at < cutoff)
            .delete(synchronize_session=False)
        )
        db.session.commit()
        if removed:
            logger.info("Removidos %s pedidos antigos", removed)
        return removed
    except Exception:
        db.session.rollback()
        logger.exception("Erro ao remover pedidos antigos")
        return 0


def normalize_items(items):
    # exige pelo menos 1 item valido
    if not isinstance(items, list) or not items:
        raise ValueError("Adicione pelo menos um item")
    normalized = []
    for idx, raw in enumerate(items, start=1):
        if not isinstance(raw, dict):
            raise ValueError(f"Item {idx} invalido")

        quantidade_raw = raw.get("quantidade")
        valor_raw = raw.get("valor")
        codigo = str(raw.get("codigo", "")).strip()
        descricao = str(raw.get("descricao", "")).strip()
        prefixo = str(raw.get("prefixo", "")).strip()
        estoque_raw = raw.get("estoque")

        empty_line = (
            (quantidade_raw in (None, "", 0, "0"))
            and not codigo
            and not descricao
            and (valor_raw in (None, "", 0, "0", 0.0))
            and not prefixo
            and (estoque_raw in (None, "", 0, "0", 0.0))
        )
        if empty_line:
            continue

        try:
            quantidade = int(quantidade_raw)
        except (TypeError, ValueError):
            raise ValueError(f"Item {idx} invalido: quantidade")
        if quantidade <= 0:
            raise ValueError(f"Item {idx} precisa de quantidade maior que zero")

        try:
            valor = float(valor_raw)
        except (TypeError, ValueError):
            raise ValueError(f"Item {idx} invalido: valor")

        if not codigo or not descricao:
            raise ValueError(f"Item {idx} precisa de codigo e descricao")

        if estoque_raw in (None, "", 0, "0", 0.0):
            estoque = None
        else:
            try:
                estoque = int(estoque_raw)
            except (TypeError, ValueError):
                raise ValueError(f"Item {idx} invalido: estoque")

        normalized.append(
            {
                "quantidade": quantidade,
                "codigo": codigo,
                "descricao": descricao,
                "prefixo": prefixo,
                "valor": valor,
                "estoque": estoque,
            }
        )

    if not normalized:
        raise ValueError("Adicione pelo menos um item valido")
    return normalized


def create_pending_order(fornecedor: str, items, creator: str | None):
    fornecedor = (fornecedor or "").strip()
    if not fornecedor:
        raise ValueError("Selecione um fornecedor")

    fornecedor_registro = Fornecedor.query.filter(
        func.upper(Fornecedor.nome) == fornecedor.upper()
    ).first()
    if not fornecedor_registro:
        raise ValueError("Fornecedor nao encontrado")

    normalized = normalize_items(items)

    pedido = Pedido(
        fornecedor=fornecedor_registro.nome,
        arquivo_excel="",
        arquivo_pdf="",
        status="Pendente",
        created_by=(creator or "").upper() or None,
    )
    db.session.add(pedido)
    db.session.flush()  # garante ID para relacionamento

    for item in normalized:
        pedido.itens.append(
            PedidoItem(
                codigo=item["codigo"],
                descricao=item["descricao"],
                quantidade=item["quantidade"],
                prefixo=item["prefixo"],
                valor=Decimal(str(item["valor"])),
                estoque=item["estoque"],
            )
        )

    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        logger.exception("Falha ao criar pedido")
        raise
    return pedido.id


def list_orders(fornecedor: str | None = None, status: str | None = None):
    query = Pedido.query
    if fornecedor:
        query = query.filter(Pedido.fornecedor.ilike(f"%{fornecedor.strip()}%"))
    if status:
        query = query.filter(Pedido.status == status)
    pedidos = query.order_by(Pedido.created_at.desc()).all()
    return [
        {
            "id": pedido.id,
            "fornecedor": pedido.fornecedor,
            "created_by": pedido.created_by,
            "created_at": pedido.created_at.isoformat() if pedido.created_at else None,
            "status": pedido.status,
            "arquivo_excel": pedido.arquivo_excel or "",
            "arquivo_pdf": pedido.arquivo_pdf or "",
        }
        for pedido in pedidos
    ]


def get_order(order_id: int):
    pedido = (
        Pedido.query.options(selectinload(Pedido.itens))
        .filter_by(id=order_id)
        .first()
    )
    if not pedido:
        return None
    itens_list = []
    for item in pedido.itens:
        valor = float(item.valor or 0)
        total = valor * item.quantidade
        itens_list.append(
            {
                "id": item.id,
                "codigo": item.codigo,
                "descricao": item.descricao,
                "quantidade": item.quantidade,
                "prefixo": item.prefixo,
                "valor": valor,
                "estoque": item.estoque,
                "total": total,
            }
        )
    return {
        "id": pedido.id,
        "fornecedor": pedido.fornecedor,
        "created_by": pedido.created_by,
        "created_at": pedido.created_at.isoformat() if pedido.created_at else None,
        "status": pedido.status,
        "arquivo_excel": pedido.arquivo_excel or "",
        "arquivo_pdf": pedido.arquivo_pdf or "",
        "itens": itens_list,
        "arquivo_excel_path": resolve_pedido_excel_path(pedido.arquivo_excel),
    }


def update_pending_order(order_id: int, items):
    normalized = normalize_items(items)
    pedido = Pedido.query.filter_by(id=order_id).first()
    if not pedido:
        raise ValueError("Pedido nao encontrado")
    if pedido.status != "Pendente":
        raise ValueError("Somente pedidos pendentes podem ser alterados")
    pedido.itens.clear()
    for item in normalized:
        pedido.itens.append(
            PedidoItem(
                codigo=item["codigo"],
                descricao=item["descricao"],
                quantidade=item["quantidade"],
                prefixo=item["prefixo"],
                valor=Decimal(str(item["valor"])),
                estoque=item["estoque"],
            )
        )
    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        logger.exception("Falha ao atualizar itens do pedido")
        raise


def approve_order(order_id: int, approver: str | None):
    pedido = Pedido.query.filter_by(id=order_id).first()
    if not pedido:
        raise ValueError("Pedido nao encontrado")
    if pedido.status != "Pendente":
        raise ValueError("Pedido ja foi processado")
    pedido.status = "Aprovado"
    if approver:
        pedido.created_by = pedido.created_by or approver
    db.session.add(pedido)
    db.session.commit()


def build_order_payload(order_id: int):
    pedido = get_order(order_id)
    if not pedido:
        raise ValueError("Pedido nao encontrado")
    created_at = pedido.get("created_at") or ""
    created_dt = None
    if created_at:
        try:
            created_dt = datetime.fromisoformat(created_at)
        except ValueError:
            try:
                created_dt = datetime.strptime(created_at.split()[0], "%Y-%m-%d")
            except Exception as exc:
                raise ValueError("Data do pedido invalida") from exc
    created_dt = created_dt or datetime.utcnow()
    payload = {
        "numero": pedido["id"],
        "fornecedor": pedido["fornecedor"],
        "data": created_dt,
        "status": pedido["status"],
        "arquivo_excel": pedido.get("arquivo_excel") or "",
        "itens": [
            {
                "codigo": item["codigo"],
                "descricao": item["descricao"],
                "quantidade": item["quantidade"],
                "prefixo": item["prefixo"],
                "valor_unitario": item["valor"],
            }
            for item in (pedido["itens"] or [])
        ],
    }
    if not payload["itens"]:
        raise ValueError("Pedido nao possui itens")
    return payload


def generate_order_file(order_id: int):
    payload = build_order_payload(order_id)
    if payload["status"] not in {"Aprovado", "Gerado"}:
        raise ValueError("Pedido precisa estar aprovado para gerar arquivo")
    caminho = gerar_arquivo_pedido_aprovado_arquivo(payload)
    pedido = Pedido.query.get(order_id)
    if not pedido:
        raise ValueError("Pedido nao encontrado")
    pedido.arquivo_excel = Path(caminho).name
    pedido.status = "Gerado"
    db.session.add(pedido)
    db.session.commit()
    return caminho


def gerar_arquivo_pedido_aprovado_arquivo(pedido_payload):
    if not MODELO_PATH.exists():
        raise FileNotFoundError(
            "Modelo modelo_pedido.xlsm nao encontrado. Verifique o caminho configurado."
        )

    fornecedor_limpo = "".join(
        c for c in pedido_payload["fornecedor"] if c.isalnum() or c in (" ", "_", "-")
    ).strip()
    fornecedor_limpo = fornecedor_limpo.replace(" ", "_") or "FORNECEDOR"
    data_str = pedido_payload["data"].strftime("%Y-%m-%d")
    nome_arquivo = f"{fornecedor_limpo}_{pedido_payload['numero']}_{data_str}.xlsx"
    caminho_saida = PASTA_PEDIDOS_APROVADOS / nome_arquivo

    try:
        shutil.copy(str(MODELO_PATH), str(caminho_saida))
    except PermissionError as exc:
        raise ValueError("O modelo esta em uso. Feche o Excel e tente novamente.") from exc
    except Exception:
        logger.exception("Erro ao copiar modelo")
        raise

    wb = load_workbook(str(caminho_saida))

    nome_aba = None
    for sheet in wb.sheetnames:
        if "IMPRESSAO" in sheet.upper():
            nome_aba = sheet
            break
    if not nome_aba:
        wb.close()
        raise ValueError("Aba de impressao nao encontrada no modelo.")

    ws = wb[nome_aba]
    try:
        ws["AG2"] = pedido_payload["numero"]
        ws["V6"] = pedido_payload["data"].day
        ws["X6"] = pedido_payload["data"].month
        ws["AG6"] = pedido_payload["data"].year
        ws["G8"] = pedido_payload["fornecedor"]
    except Exception as exc:
        wb.close()
        raise ValueError(f"Erro ao preencher campos do modelo: {exc}") from exc

    linha_inicial = 15
    for index, item in enumerate(pedido_payload["itens"]):
        linha = linha_inicial + index
        ws[f"A{linha}"] = item["quantidade"]
        ws[f"F{linha}"] = item["prefixo"]
        ws[f"J{linha}"] = item["codigo"]
        ws[f"K{linha}"] = item["descricao"]
        ws[f"X{linha}"] = item["valor_unitario"]

    wb.save(str(caminho_saida))
    wb.close()
    return str(caminho_saida)


def resolve_pedido_excel_path(arquivo_excel: str | None):
    if not arquivo_excel:
        return None
    candidate = Path(arquivo_excel)
    if candidate.is_absolute() and candidate.exists():
        return str(candidate)

    filename = candidate.name
    for base_dir in (
        PASTA_PEDIDOS_APROVADOS,
        PASTA_PEDIDOS_GERADOS,
        STORAGE_ROOT,
    ):
        resolved = (Path(base_dir) / filename).resolve()
        if resolved.exists():
            return str(resolved)

    if candidate.is_absolute():
        return str(candidate)
    return str((PASTA_PEDIDOS_APROVADOS / filename).resolve())


# =====================================================
# ROTAS / VIEWS
# =====================================================
def _render_login(template_name: str):
    error = None
    if request.method == "POST":
        username = (request.form.get("username") or "").strip().upper()
        password = (request.form.get("password") or "").strip()
        user = verificar_login(username, password)
        if user:
            session.clear()
            session.permanent = True
            session["user"] = user
            next_url = request.args.get("next") or url_for("dashboard")
            return redirect(next_url)
        error = "Usuario ou senha invalidos"

    if "user" in session and error is None:
        return redirect(url_for("dashboard"))
    return render_template(template_name, error=error)


@app.route("/login", methods=["GET", "POST"])
def login():
    return _render_login("login.html")


@app.route("/login/compact", methods=["GET", "POST"])
def login_compact():
    return _render_login("login_card.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/")
@login_required
def dashboard():
    return render_template("dashboard.html", user=current_user())


@app.route("/dashboard/compact")
@login_required
def dashboard_compact():
    return render_template("dashboard_compact.html", user=current_user())


@app.route("/api/context")
@api_login_required
def api_context():
    user = current_user()
    data = {
        "user": user,
        "suppliers": list_suppliers(),
        "statuses": ["Pendente", "Aprovado", "Gerado"],
        "purged_on_start": PURGED_ON_START,
        "modelo_disponivel": MODELO_PATH.exists(),
    }
    if user and user.get("role") == "admin":
        data["users"] = list_users()
    return jsonify(data)


@app.route("/api/fornecedores", methods=["GET", "POST"])
@api_login_required
def api_fornecedores():
    if request.method == "GET":
        return jsonify({"suppliers": list_suppliers()})

    payload = request.get_json(silent=True) or {}
    nome = payload.get("nome") or payload.get("name")
    try:
        add_supplier(nome)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    return jsonify({"suppliers": list_suppliers()})


@app.route("/api/pedidos", methods=["GET", "POST"])
@api_login_required
def api_pedidos():
    if request.method == "GET":
        fornecedor = (request.args.get("fornecedor") or "").strip()
        status = (request.args.get("status") or "").strip()
        pedidos = list_orders(fornecedor or None, status or None)
        return jsonify({"pedidos": pedidos})

    payload = request.get_json(silent=True) or {}
    fornecedor = payload.get("fornecedor")
    itens = payload.get("itens") or []
    try:
        pedido_id = create_pending_order(
            fornecedor,
            itens,
            current_user().get("username") if current_user() else None,
        )
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception:
        traceback.print_exc()
        return jsonify({"error": "Falha ao criar pedido"}), 500
    return jsonify({"pedido_id": pedido_id})


@app.route("/api/pedidos/<int:pedido_id>", methods=["GET", "PUT"])
@api_login_required
def api_pedido_detalhe(pedido_id):
    if request.method == "GET":
        pedido = get_order(pedido_id)
        if not pedido:
            return jsonify({"error": "Pedido nao encontrado"}), 404
        return jsonify({"pedido": pedido})

    payload = request.get_json(silent=True) or {}
    itens = payload.get("itens") or []
    try:
        update_pending_order(pedido_id, itens)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception:
        traceback.print_exc()
        return jsonify({"error": "Falha ao atualizar pedido"}), 500
    pedido = get_order(pedido_id)
    return jsonify({"pedido": pedido})


@app.route("/api/pedidos/<int:pedido_id>/approve", methods=["POST"])
@roles_required("approver", "admin")
def api_pedido_approve(pedido_id):
    try:
        approve_order(pedido_id, current_user().get("username"))
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception:
        traceback.print_exc()
        return jsonify({"error": "Falha ao aprovar pedido"}), 500

    download_url = None
    try:
        caminho = generate_order_file(pedido_id)
        download_url = url_for("download_pedido", pedido_id=pedido_id)
    except Exception as exc:
        traceback.print_exc()
        return jsonify(
            {
                "warning": "Pedido aprovado, mas houve erro ao gerar arquivo automaticamente.",
                "detail": str(exc),
            }
        )

    pedido = get_order(pedido_id)
    return jsonify(
        {
            "pedido": pedido,
            "download_url": download_url,
            "caminho": caminho,
        }
    )


@app.route("/api/pedidos/<int:pedido_id>/generate", methods=["POST"])
@api_login_required
def api_pedido_generate(pedido_id):
    try:
        caminho = generate_order_file(pedido_id)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception:
        traceback.print_exc()
        return jsonify({"error": "Falha ao gerar arquivo"}), 500
    pedido = get_order(pedido_id)
    return jsonify(
        {
            "pedido": pedido,
            "download_url": url_for("download_pedido", pedido_id=pedido_id),
            "caminho": caminho,
        }
    )


@app.route("/pedidos/<int:pedido_id>/download")
@login_required
def download_pedido(pedido_id):
    pedido = get_order(pedido_id)
    if not pedido:
        abort(404)
    caminho = pedido.get("arquivo_excel_path") or resolve_pedido_excel_path(
        pedido.get("arquivo_excel")
    )
    if not caminho or not os.path.exists(caminho):
        abort(404)
    nome_arquivo = os.path.basename(caminho)
    return send_file(caminho, as_attachment=True, download_name=nome_arquivo)


@app.route("/pedidos/<int:pedido_id>")
@login_required
def pedido_detalhe_pagina(pedido_id):
    pedido = get_order(pedido_id)
    if not pedido:
        abort(404)
    return render_template("order_detail.html", pedido=pedido, user=current_user())


@app.route("/api/users", methods=["GET", "POST"])
@roles_required("admin")
def api_users():
    if request.method == "GET":
        return jsonify({"users": list_users()})

    payload = request.get_json(silent=True) or {}
    try:
        create_user(
            payload.get("username"),
            payload.get("password"),
            payload.get("role") or "creator",
        )
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    return jsonify({"users": list_users()})


@app.route("/api/users/<username>", methods=["DELETE"])
@roles_required("admin")
def api_users_delete(username):
    username = (username or "").strip().upper()
    if username == (current_user() or {}).get("username"):
        return jsonify({"error": "Nao e possivel remover o proprio usuario"}), 400
    try:
        delete_user(username)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 404
    return jsonify({"users": list_users()})


@app.route("/api/me/password", methods=["POST"])
@api_login_required
def api_me_password():
    payload = request.get_json(silent=True) or {}
    nova = (payload.get("new_password") or "").strip()
    if not nova:
        return jsonify({"error": "Informe a nova senha"}), 400
    try:
        change_password(current_user()["username"], nova)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    return jsonify({"status": "ok"})


@app.route("/health")
def health():
    try:
        db.session.execute(text("SELECT 1"))
        return jsonify({"status": "ok"})
    except Exception:
        traceback.print_exc()
        return jsonify({"status": "error"}), 500


# =====================================================
# INIT
# =====================================================
PURGED_ON_START = 0


def _initialize_app():
    global PURGED_ON_START
    with app.app_context():
        # cria as tabelas automaticamente no banco (Postgres) e garante usuarios
        db.create_all()
        garantir_usuarios_iniciais()
        PURGED_ON_START = purge_old_pedidos()


_initialize_app()


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    debug = os.environ.get("FLASK_DEBUG", "").lower() in {"1", "true", "yes"}
    app.run(host="0.0.0.0", port=port, debug=debug)

