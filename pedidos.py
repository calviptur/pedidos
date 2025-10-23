import os
import sqlite3
import shutil
import traceback
from datetime import datetime, timedelta
from functools import wraps
from pathlib import Path

from flask import (
    Flask,
    abort,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from openpyxl import load_workbook


ROOT_DIR = Path(__file__).resolve().parent
STORAGE_ROOT = Path(
    os.environ.get("PEDIDOS_STORAGE_DIR")
    or os.environ.get("PEDIDOS_BASE_DIR")
    or ROOT_DIR
).expanduser()
STORAGE_ROOT.mkdir(parents=True, exist_ok=True)

DB_PATH = Path(os.environ.get("PEDIDOS_DB_PATH") or (STORAGE_ROOT / "pedidos.db"))
if not DB_PATH.exists():
    legacy_db = ROOT_DIR.parent / "pedidos.db"
    if legacy_db.exists():
        DB_PATH = legacy_db

modelo_default = ROOT_DIR / "modelo_pedido.xlsm"
MODELO_PATH = Path(os.environ.get("PEDIDOS_MODELO_PATH") or modelo_default)
if not MODELO_PATH.exists():
    alt_modelo = STORAGE_ROOT / "modelo_pedido.xlsm"
    if alt_modelo.exists():
        MODELO_PATH = alt_modelo
    else:
        legacy_modelo = ROOT_DIR.parent / "modelo_pedido.xlsm"
        if legacy_modelo.exists():
            MODELO_PATH = legacy_modelo

PASTA_PEDIDOS_GERADOS = Path(
    os.environ.get("PEDIDOS_GERADOS_DIR") or (STORAGE_ROOT / "Pedidos Gerados")
)
PASTA_PEDIDOS_APROVADOS = Path(
    os.environ.get("PEDIDOS_APROVADOS_DIR") or (STORAGE_ROOT / "PedidosAprovados")
)

if not os.environ.get("PEDIDOS_GERADOS_DIR"):
    legacy_gerados = ROOT_DIR.parent / "Pedidos Gerados"
    if legacy_gerados.exists():
        PASTA_PEDIDOS_GERADOS = legacy_gerados

if not os.environ.get("PEDIDOS_APROVADOS_DIR"):
    legacy_aprovados = ROOT_DIR.parent / "PedidosAprovados"
    if legacy_aprovados.exists():
        PASTA_PEDIDOS_APROVADOS = legacy_aprovados

PASTA_PEDIDOS_GERADOS.mkdir(parents=True, exist_ok=True)
PASTA_PEDIDOS_APROVADOS.mkdir(parents=True, exist_ok=True)

app = Flask(__name__, template_folder="templates", static_folder="static")
app.config["SECRET_KEY"] = os.environ.get("PEDIDOS_SECRET_KEY", "change-me")
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(hours=8)


# =========================
# Database helpers
# =========================
def conectar_db():
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS pedidos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fornecedor TEXT,
            arquivo_excel TEXT,
            arquivo_pdf TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            status TEXT DEFAULT 'Pendente',
            created_by TEXT
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS fornecedores (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT UNIQUE
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE,
            password TEXT,
            role TEXT
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS pedidos_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            pedido_id INTEGER,
            codigo TEXT,
            descricao TEXT,
            quantidade INTEGER,
            prefixo TEXT,
            valor REAL,
            estoque INTEGER,
            FOREIGN KEY(pedido_id) REFERENCES pedidos(id)
        )
        """
    )
    conn.commit()
    return conn, cur


def garantir_usuarios_iniciais():
    conn, cur = conectar_db()
    defaults = [
        ("MIGUEL", "1234", "creator"),
        ("MICHEL", "1234", "approver"),
        ("LUCAS", "1234", "admin"),
    ]
    for username, password, role in defaults:
        cur.execute("SELECT id FROM users WHERE username=?", (username,))
        if not cur.fetchone():
            cur.execute(
                "INSERT INTO users (username, password, role) VALUES (?, ?, ?)",
                (username, password, role),
            )
    conn.commit()
    conn.close()


def purge_old_pedidos():
    conn, cur = conectar_db()
    cutoff = datetime.now() - timedelta(days=135)
    cutoff_str = cutoff.strftime("%Y-%m-%d %H:%M:%S")
    cur.execute("DELETE FROM pedidos WHERE created_at < ?", (cutoff_str,))
    removed = cur.rowcount
    conn.commit()
    conn.close()
    return removed


# =========================
# User management
# =========================
def fetch_user(username):
    conn, cur = conectar_db()
    cur.execute(
        "SELECT id, username, password, role FROM users WHERE username=?",
        (username,),
    )
    row = cur.fetchone()
    conn.close()
    return dict(row) if row else None


def verificar_login(username, password):
    user = fetch_user(username)
    if user and user["password"] == password:
        return user
    return None


def change_password(username, new_password):
    conn, cur = conectar_db()
    cur.execute("UPDATE users SET password=? WHERE username=?", (new_password, username))
    conn.commit()
    updated = cur.rowcount
    conn.close()
    if not updated:
        raise ValueError("Usuario nao encontrado")


def create_user(username, password, role):
    username = (username or "").strip()
    password = (password or "").strip()
    role = (role or "creator").strip()
    if not username or not password:
        raise ValueError("Usuario e senha devem ser informados")

    conn, cur = conectar_db()
    try:
        cur.execute(
            "INSERT INTO users (username, password, role) VALUES (?, ?, ?)",
            (username.upper(), password, role),
        )
        conn.commit()
    except sqlite3.IntegrityError as exc:
        raise ValueError("Usuario ja existe") from exc
    finally:
        conn.close()


def delete_user(username):
    conn, cur = conectar_db()
    cur.execute("DELETE FROM users WHERE username=?", (username,))
    conn.commit()
    removed = cur.rowcount
    conn.close()
    if not removed:
        raise ValueError("Usuario nao encontrado")


def list_users():
    conn, cur = conectar_db()
    cur.execute("SELECT username, role FROM users ORDER BY username")
    users = [dict(row) for row in cur.fetchall()]
    conn.close()
    return users


# =========================
# Fornecedores
# =========================
def add_supplier(nome):
    nome = (nome or "").strip()
    if not nome:
        raise ValueError("Informe o nome do fornecedor")
    conn, cur = conectar_db()
    try:
        cur.execute("INSERT INTO fornecedores (nome) VALUES (?)", (nome,))
        conn.commit()
    except sqlite3.IntegrityError as exc:
        raise ValueError("Fornecedor ja existe") from exc
    finally:
        conn.close()
    return nome


def list_suppliers():
    conn, cur = conectar_db()
    cur.execute("SELECT nome FROM fornecedores ORDER BY nome")
    fornecedores = [row["nome"] for row in cur.fetchall()]
    conn.close()
    return fornecedores


# =========================
# Pedidos helpers
# =========================
def normalize_items(items):
    if not isinstance(items, list) or not items:
        raise ValueError("Adicione pelo menos um item")
    normalized = []
    for idx, raw in enumerate(items, start=1):
        try:
            quantidade = int(raw.get("quantidade"))
            codigo = str(raw.get("codigo", "")).strip()
            descricao = str(raw.get("descricao", "")).strip()
            prefixo = str(raw.get("prefixo", "")).strip()
            valor = float(raw.get("valor"))
            estoque = int(raw.get("estoque"))
        except (TypeError, ValueError):
            raise ValueError(f"Item {idx} invalido")
        if quantidade <= 0:
            raise ValueError(f"Item {idx} precisa de quantidade maior que zero")
        if not codigo or not descricao:
            raise ValueError(f"Item {idx} precisa de codigo e descricao")
        normalized.append(
            {
                "quantidade": quantidade,
                "codigo": codigo,
                "descricao": descricao,
                "prefixo": prefixo,
                "valor": valor,
                "estoque": estoque,
            }
        )
    return normalized


def create_pending_order(fornecedor, items, creator):
    fornecedor = (fornecedor or "").strip()
    if not fornecedor:
        raise ValueError("Selecione um fornecedor")
    normalized = normalize_items(items)

    conn, cur = conectar_db()
    cur.execute("SELECT id FROM fornecedores WHERE nome=?", (fornecedor,))
    if not cur.fetchone():
        conn.close()
        raise ValueError("Fornecedor nao encontrado")

    cur.execute(
        "INSERT INTO pedidos (fornecedor, arquivo_excel, arquivo_pdf, status, created_by) VALUES (?, ?, ?, ?, ?)",
        (fornecedor, "", "", "Pendente", creator or ""),
    )
    pedido_id = cur.lastrowid
    for item in normalized:
        cur.execute(
            """
            INSERT INTO pedidos_items (pedido_id, codigo, descricao, quantidade, prefixo, valor, estoque)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                pedido_id,
                item["codigo"],
                item["descricao"],
                item["quantidade"],
                item["prefixo"],
                item["valor"],
                item["estoque"],
            ),
        )
    conn.commit()
    conn.close()
    return pedido_id


def list_orders(fornecedor=None, status=None):
    conn, cur = conectar_db()
    query = "SELECT id, fornecedor, created_by, created_at, status, arquivo_excel, arquivo_pdf FROM pedidos"
    clauses = []
    params = []
    if fornecedor:
        clauses.append("fornecedor LIKE ?")
        params.append(f"%{fornecedor}%")
    if status:
        clauses.append("status = ?")
        params.append(status)
    if clauses:
        query += " WHERE " + " AND ".join(clauses)
    query += " ORDER BY created_at DESC"
    cur.execute(query, params)
    pedidos = [dict(row) for row in cur.fetchall()]
    conn.close()
    return pedidos


def get_order(order_id):
    conn, cur = conectar_db()
    cur.execute(
        "SELECT id, fornecedor, created_by, created_at, status, arquivo_excel, arquivo_pdf FROM pedidos WHERE id=?",
        (order_id,),
    )
    pedido = cur.fetchone()
    if not pedido:
        conn.close()
        return None

    cur.execute(
        """
        SELECT id, codigo, descricao, quantidade, prefixo, valor, estoque
        FROM pedidos_items
        WHERE pedido_id=?
        """,
        (order_id,),
    )
    itens = [
        {
            "id": row["id"],
            "codigo": row["codigo"],
            "descricao": row["descricao"],
            "quantidade": row["quantidade"],
            "prefixo": row["prefixo"],
            "valor": row["valor"],
            "estoque": row["estoque"],
            "total": row["quantidade"] * row["valor"],
        }
        for row in cur.fetchall()
    ]
    conn.close()
    pedido_dict = dict(pedido)
    pedido_dict["itens"] = itens
    return pedido_dict


def update_pending_order(order_id, items):
    normalized = normalize_items(items)
    conn, cur = conectar_db()
    cur.execute("SELECT status FROM pedidos WHERE id=?", (order_id,))
    row = cur.fetchone()
    if not row:
        conn.close()
        raise ValueError("Pedido nao encontrado")
    if row["status"] != "Pendente":
        conn.close()
        raise ValueError("Somente pedidos pendentes podem ser alterados")

    cur.execute("DELETE FROM pedidos_items WHERE pedido_id=?", (order_id,))
    for item in normalized:
        cur.execute(
            """
            INSERT INTO pedidos_items (pedido_id, codigo, descricao, quantidade, prefixo, valor, estoque)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                order_id,
                item["codigo"],
                item["descricao"],
                item["quantidade"],
                item["prefixo"],
                item["valor"],
                item["estoque"],
            ),
        )
    conn.commit()
    conn.close()


def approve_order(order_id, approver):
    conn, cur = conectar_db()
    cur.execute("SELECT status FROM pedidos WHERE id=?", (order_id,))
    row = cur.fetchone()
    if not row:
        conn.close()
        raise ValueError("Pedido nao encontrado")
    if row["status"] != "Pendente":
        conn.close()
        raise ValueError("Pedido ja foi processado")
    cur.execute("UPDATE pedidos SET status='Aprovado' WHERE id=?", (order_id,))
    conn.commit()
    conn.close()


def build_order_payload(order_id):
    pedido = get_order(order_id)
    if not pedido:
        raise ValueError("Pedido nao encontrado")
    created_at = pedido.get("created_at") or ""
    try:
        created_dt = datetime.fromisoformat(created_at)
    except ValueError:
        try:
            created_dt = datetime.strptime(created_at.split()[0], "%Y-%m-%d")
        except Exception as exc:
            raise ValueError("Data do pedido invalida") from exc
    payload = {
        "numero": pedido["id"],
        "fornecedor": pedido["fornecedor"],
        "data": created_dt,
        "status": pedido["status"],
        "arquivo_excel": pedido.get("arquivo_excel") or "",
        "itens": [
            {
                "codigo": item["codigo"],
                "descricao": item["descricao"],
                "quantidade": item["quantidade"],
                "prefixo": item["prefixo"],
                "valor_unitario": item["valor"],
            }
            for item in pedido["itens"]
        ],
    }
    if not payload["itens"]:
        raise ValueError("Pedido nao possui itens")
    return payload


def generate_order_file(order_id):
    payload = build_order_payload(order_id)
    if payload["status"] not in {"Aprovado", "Gerado"}:
        raise ValueError("Pedido precisa estar aprovado para gerar arquivo")
    caminho = gerar_arquivo_pedido_aprovado_arquivo(payload)
    conn, cur = conectar_db()
    cur.execute(
        "UPDATE pedidos SET arquivo_excel=?, status='Gerado' WHERE id=?",
        (caminho, order_id),
    )
    conn.commit()
    conn.close()
    return caminho


def gerar_arquivo_pedido_aprovado_arquivo(pedido):
    if not MODELO_PATH.exists():
        raise FileNotFoundError(
            "Modelo modelo_pedido.xlsm nao encontrado. Verifique o caminho configurado."
        )

    fornecedor_limpo = "".join(
        c for c in pedido["fornecedor"] if c.isalnum() or c in (" ", "_", "-")
    ).strip()
    fornecedor_limpo = fornecedor_limpo.replace(" ", "_") or "FORNECEDOR"
    data_str = pedido["data"].strftime("%Y-%m-%d")
    nome_arquivo = f"{fornecedor_limpo}_{data_str}.xlsx"
    caminho_saida = PASTA_PEDIDOS_APROVADOS / nome_arquivo

    shutil.copy(str(MODELO_PATH), str(caminho_saida))
    wb = load_workbook(str(caminho_saida))

    nome_aba = None
    for sheet in wb.sheetnames:
        if "IMPRESSAO" in sheet.upper():
            nome_aba = sheet
            break
    if not nome_aba:
        wb.close()
        raise ValueError("Aba de impressao nao encontrada no modelo.")

    ws = wb[nome_aba]
    ws["AG2"] = pedido["numero"]
    ws["V6"] = pedido["data"].day
    ws["X6"] = pedido["data"].month
    ws["AG6"] = pedido["data"].year
    ws["G8"] = pedido["fornecedor"]

    linha_inicial = 15
    for index, item in enumerate(pedido["itens"]):
        linha = linha_inicial + index
        ws[f"A{linha}"] = item["quantidade"]
        ws[f"F{linha}"] = item["prefixo"]
        ws[f"J{linha}"] = item["codigo"]
        ws[f"K{linha}"] = item["descricao"]
        ws[f"X{linha}"] = item["valor_unitario"]

    wb.save(str(caminho_saida))
    wb.close()
    return str(caminho_saida)


# prepare database defaults on import
garantir_usuarios_iniciais()
PURGED_ON_START = purge_old_pedidos()


# =========================
# Authentication helpers
# =========================
def current_user():
    return session.get("user")


def login_required(view):
    @wraps(view)
    def wrapper(*args, **kwargs):
        if "user" not in session:
            next_url = request.path if request.method == "GET" else None
            return redirect(url_for("login", next=next_url))
        return view(*args, **kwargs)

    return wrapper


def api_login_required(view):
    @wraps(view)
    def wrapper(*args, **kwargs):
        if "user" not in session:
            return jsonify({"error": "autenticacao requerida"}), 401
        return view(*args, **kwargs)

    return wrapper


def roles_required(*roles):
    def decorator(view):
        @wraps(view)
        def wrapper(*args, **kwargs):
            user = current_user()
            if not user:
                return jsonify({"error": "autenticacao requerida"}), 401
            if user["role"] not in roles:
                return jsonify({"error": "permissao negada"}), 403
            return view(*args, **kwargs)

        return wrapper

    return decorator


# =========================
# Routes
# =========================
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = (request.form.get("username") or "").strip().upper()
        password = (request.form.get("password") or "").strip()
        user = verificar_login(username, password)
        if user:
            session.clear()
            session.permanent = True
            session["user"] = {
                "id": user["id"],
                "username": user["username"],
                "role": user["role"],
            }
            next_url = request.args.get("next") or url_for("dashboard")
            return redirect(next_url)
        return render_template("login.html", error="Usuario ou senha invalidos")

    if "user" in session:
        return redirect(url_for("dashboard"))
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/")
@login_required
def dashboard():
    return render_template("dashboard.html", user=current_user())


@app.route("/api/context")
@api_login_required
def api_context():
    user = current_user()
    data = {
        "user": user,
        "suppliers": list_suppliers(),
        "statuses": ["Pendente", "Aprovado", "Gerado"],
        "purged_on_start": PURGED_ON_START,
        "modelo_disponivel": MODELO_PATH.exists(),
    }
    if user["role"] == "admin":
        data["users"] = list_users()
    return jsonify(data)


@app.route("/api/fornecedores", methods=["GET", "POST"])
@api_login_required
def api_fornecedores():
    if request.method == "GET":
        return jsonify({"suppliers": list_suppliers()})

    payload = request.get_json(silent=True) or {}
    nome = payload.get("nome") or payload.get("name")
    try:
        novo = add_supplier(nome)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    return jsonify({"supplier": novo, "suppliers": list_suppliers()})


@app.route("/api/pedidos", methods=["GET", "POST"])
@api_login_required
def api_pedidos():
    if request.method == "GET":
        fornecedor = (request.args.get("fornecedor") or "").strip()
        status = (request.args.get("status") or "").strip()
        pedidos = list_orders(fornecedor or None, status or None)
        return jsonify({"pedidos": pedidos})

    payload = request.get_json(silent=True) or {}
    fornecedor = payload.get("fornecedor")
    itens = payload.get("itens") or []
    try:
        pedido_id = create_pending_order(
            fornecedor, itens, current_user().get("username")
        )
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception:
        traceback.print_exc()
        return jsonify({"error": "Falha ao criar pedido"}), 500
    return jsonify({"pedido_id": pedido_id})


@app.route("/api/pedidos/<int:pedido_id>", methods=["GET", "PUT"])
@api_login_required
def api_pedido_detalhe(pedido_id):
    if request.method == "GET":
        pedido = get_order(pedido_id)
        if not pedido:
            return jsonify({"error": "Pedido nao encontrado"}), 404
        return jsonify({"pedido": pedido})

    payload = request.get_json(silent=True) or {}
    itens = payload.get("itens") or []
    try:
        update_pending_order(pedido_id, itens)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception:
        traceback.print_exc()
        return jsonify({"error": "Falha ao atualizar pedido"}), 500
    pedido = get_order(pedido_id)
    return jsonify({"pedido": pedido})


@app.route("/api/pedidos/<int:pedido_id>/approve", methods=["POST"])
@roles_required("approver", "admin")
def api_pedido_approve(pedido_id):
    try:
        approve_order(pedido_id, current_user().get("username"))
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception:
        traceback.print_exc()
        return jsonify({"error": "Falha ao aprovar pedido"}), 500

    download_url = None
    try:
        caminho = generate_order_file(pedido_id)
        download_url = url_for("download_pedido", pedido_id=pedido_id)
    except Exception as exc:
        traceback.print_exc()
        return jsonify(
            {
                "warning": "Pedido aprovado, mas houve erro ao gerar arquivo automaticamente.",
                "detail": str(exc),
            }
        )

    pedido = get_order(pedido_id)
    return jsonify(
        {
            "pedido": pedido,
            "download_url": download_url,
        }
    )


@app.route("/api/pedidos/<int:pedido_id>/generate", methods=["POST"])
@api_login_required
def api_pedido_generate(pedido_id):
    try:
        caminho = generate_order_file(pedido_id)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception:
        traceback.print_exc()
        return jsonify({"error": "Falha ao gerar arquivo"}), 500
    pedido = get_order(pedido_id)
    return jsonify(
        {
            "pedido": pedido,
            "download_url": url_for("download_pedido", pedido_id=pedido_id),
            "caminho": caminho,
        }
    )


@app.route("/pedidos/<int:pedido_id>/download")
@login_required
def download_pedido(pedido_id):
    pedido = get_order(pedido_id)
    if not pedido:
        abort(404)
    caminho = pedido.get("arquivo_excel")
    if not caminho or not os.path.exists(caminho):
        abort(404)
    nome_arquivo = os.path.basename(caminho)
    return send_file(caminho, as_attachment=True, download_name=nome_arquivo)


@app.route("/api/users", methods=["GET", "POST"])
@roles_required("admin")
def api_users():
    if request.method == "GET":
        return jsonify({"users": list_users()})

    payload = request.get_json(silent=True) or {}
    try:
        create_user(
            payload.get("username"),
            payload.get("password"),
            payload.get("role") or "creator",
        )
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    return jsonify({"users": list_users()})


@app.route("/api/users/<username>", methods=["DELETE"])
@roles_required("admin")
def api_users_delete(username):
    username = username.upper()
    if username == current_user().get("username"):
        return jsonify({"error": "Nao e possivel remover o proprio usuario"}), 400
    try:
        delete_user(username)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 404
    return jsonify({"users": list_users()})


@app.route("/api/me/password", methods=["POST"])
@api_login_required
def api_me_password():
    payload = request.get_json(silent=True) or {}
    nova = (payload.get("new_password") or "").strip()
    if not nova:
        return jsonify({"error": "Informe a nova senha"}), 400
    try:
        change_password(current_user()["username"], nova)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    return jsonify({"status": "ok"})


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    debug = os.environ.get("FLASK_DEBUG", "").lower() in {"1", "true", "yes"}
    app.run(host="0.0.0.0", port=port, debug=debug)
